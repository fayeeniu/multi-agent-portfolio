"""Offline company research-case intake and review contracts.

This module never performs network retrieval or model inference. Submitted identifiers,
domains, names, and documents remain claims until a named reviewer records a decision.
"""

from __future__ import annotations

import csv
import io
import ipaddress
import os
import re
import zipfile
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, sessionmaker

from .enums import (
    CompanyRelationshipType,
    CompanyEntityType,
    CompanyLifecycleStatus,
    DataClassification,
    EvidenceScope,
    IdentifierScheme,
    IdentityCandidateStatus,
    IdentityDecisionType,
    IntakeArtifactKind,
    LinkReviewStatus,
    ResearchCaseStatus,
    ResolutionStatus,
)
from .identity import (
    identifier_review_projection,
    identifier_source_key,
    is_valid_companies_house_number,
    normalize_company_name,
    normalize_identifier,
    synchronize_company_identity_review_state,
)
from .ids import sha256_bytes, stable_hash
from .models import (
    CompanyDomainDecisionModel,
    CompanyDomainModel,
    CompanyIdentifierDecisionModel,
    CompanyIdentifierModel,
    CompanyModel,
    CompanyRelationshipDecisionModel,
    CompanyRelationshipModel,
    IntakeArtifactModel,
    ResearchCaseModel,
    ResearchTemplateModel,
    ResearchTemplateVersionModel,
)

MAX_INTAKE_BYTES = 20 * 1024 * 1024
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 500
MAX_ARCHIVE_DEPTH = 10
MAX_ARCHIVE_COMPRESSION_RATIO = 100
MAX_BULK_ROWS = 10_000
MAX_BULK_COLUMNS = 32
CORE_TEMPLATE_KEY = "core_company_profile"
CORE_TEMPLATE_VERSION = "1.0.0"
INTAKE_CONTRACT_VERSION = "company-intake-v1"
HYBRID_DOCUMENT_CONTRACT_VERSION = "company-document-batch-v1"
MAX_DOCUMENTS_PER_BATCH = 12

_CORE_TEMPLATE_CONTRACT: dict[str, Any] = {
    "key": CORE_TEMPLATE_KEY,
    "version": CORE_TEMPLATE_VERSION,
    "objective": (
        "Organise company-level legal identity, verified first-party domain claims, documents, "
        "coverage, contradictions, and limitations without live retrieval or personal profiling."
    ),
    "required_capabilities": ["identity_review"],
    "optional_capabilities": [
        "companies_house_fixture",
        "first_party_domain_claim",
        "authorised_document",
    ],
    "claim_keys": [
        "legal_identity",
        "entity_status",
        "verified_domain",
        "document_inventory",
    ],
    "budgets": {"network_requests": 0, "model_tokens": 0, "background_tasks": 0},
}
CORE_TEMPLATE_SHA256 = stable_hash(_CORE_TEMPLATE_CONTRACT)

_DOCUMENT_TYPES: dict[str, tuple[str, ...]] = {
    ".pdf": ("application/pdf",),
    ".docx": ("application/vnd.openxmlformats-officedocument.wordprocessingml.document",),
    ".xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",),
    ".csv": ("text/csv", "application/csv", "text/plain"),
    ".json": ("application/json", "text/json", "text/plain"),
    ".html": ("text/html", "application/xhtml+xml", "text/plain"),
    ".htm": ("text/html", "application/xhtml+xml", "text/plain"),
    ".xml": ("application/xml", "text/xml", "text/plain"),
    ".txt": ("text/plain",),
    ".png": ("image/png",),
    ".jpg": ("image/jpeg",),
    ".jpeg": ("image/jpeg",),
}


class CompanyIntakeValidationError(ValueError):
    """The submitted company-intake contract is invalid or unsafe."""


@dataclass(frozen=True, slots=True)
class CompanyIntakeRequest:
    actor: str
    purpose: str
    classification: DataClassification
    companies_house_number: str | None = None
    website: str | None = None
    company_name: str | None = None
    jurisdiction: str | None = None
    document_bytes: bytes | None = None
    document_filename: str | None = None
    declared_mime: str | None = None


@dataclass(frozen=True, slots=True)
class CompanyIntakeResult:
    company_id: str
    research_case_id: str
    artifact_id: str
    reused_existing: bool


@dataclass(frozen=True, slots=True)
class CompanyDocumentUpload:
    content: bytes
    filename: str
    declared_mime: str | None = None


@dataclass(frozen=True, slots=True)
class _PreparedIntake:
    kind: IntakeArtifactKind
    actor: str
    purpose: str
    classification: DataClassification
    companies_house_number: str | None
    website: str | None
    domain: str | None
    company_name: str | None
    jurisdiction: str | None
    document_bytes: bytes | None
    document_filename: str | None
    declared_mime: str | None
    content_sha256: str | None
    submitted_value: dict[str, Any]
    fingerprint: str


def _clean_actor_and_purpose(actor: str, purpose: str) -> tuple[str, str]:
    clean_actor = actor.strip()
    clean_purpose = " ".join(purpose.split())
    if len(clean_actor) < 2:
        raise CompanyIntakeValidationError("Company intake requires a named actor.")
    if len(clean_purpose) < 5:
        raise CompanyIntakeValidationError("Company intake requires a substantive purpose.")
    return clean_actor, clean_purpose


def _safe_filename(filename: str) -> str:
    basename = Path(filename).name
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", basename).strip("._")
    return safe or "document.bin"


def _normalise_jurisdiction(value: str | None) -> str | None:
    if value is None or not value.strip():
        return None
    normalized = " ".join(value.strip().upper().split())
    if len(normalized) > 64 or re.fullmatch(r"[A-Z0-9][A-Z0-9 .'-]*", normalized) is None:
        raise CompanyIntakeValidationError("Jurisdiction must be a short declared code or name.")
    return normalized


def _normalise_website(value: str | None) -> tuple[str | None, str | None]:
    if value is None or not value.strip():
        return None, None
    parsed = urlsplit(value.strip())
    if parsed.scheme.lower() != "https":
        raise CompanyIntakeValidationError("Company website claims must use HTTPS.")
    if parsed.username is not None or parsed.password is not None:
        raise CompanyIntakeValidationError("Company website claims must not contain credentials.")
    if parsed.hostname is None:
        raise CompanyIntakeValidationError("Company website claim has no hostname.")
    try:
        domain = parsed.hostname.rstrip(".").encode("idna").decode("ascii").lower()
    except UnicodeError as exc:
        raise CompanyIntakeValidationError("Company website hostname is invalid.") from exc
    try:
        ipaddress.ip_address(domain)
    except ValueError:
        pass
    else:
        raise CompanyIntakeValidationError("Company website must use a public hostname, not an IP.")
    if domain == "localhost" or domain.endswith(".localhost") or "." not in domain:
        raise CompanyIntakeValidationError("Company website must use a public hostname.")
    try:
        port = parsed.port
    except ValueError as exc:
        raise CompanyIntakeValidationError("Company website port is invalid.") from exc
    if port not in {None, 443}:
        raise CompanyIntakeValidationError("Company website claim may use only the HTTPS port.")
    path = re.sub(r"/{2,}", "/", parsed.path or "/")
    if path != "/":
        path = path.rstrip("/")
    canonical = urlunsplit(("https", domain, path, "", ""))
    return canonical, domain


def _validate_zip_document(payload: bytes, suffix: str) -> None:
    try:
        archive = zipfile.ZipFile(io.BytesIO(payload))
    except zipfile.BadZipFile as exc:
        raise CompanyIntakeValidationError(
            f"{suffix} document is not a valid ZIP container."
        ) from exc
    with archive:
        entries = archive.infolist()
        if len(entries) > MAX_ARCHIVE_ENTRIES:
            raise CompanyIntakeValidationError("Document archive contains too many entries.")
        total_uncompressed = 0
        names: set[str] = set()
        for entry in entries:
            normalized = entry.filename.replace("\\", "/")
            parts = tuple(part for part in normalized.split("/") if part)
            if normalized.startswith("/") or ".." in parts:
                raise CompanyIntakeValidationError("Document archive contains an unsafe path.")
            if len(parts) > MAX_ARCHIVE_DEPTH:
                raise CompanyIntakeValidationError("Document archive nesting is too deep.")
            if entry.flag_bits & 0x1:
                raise CompanyIntakeValidationError("Encrypted document archives are unsupported.")
            total_uncompressed += entry.file_size
            if total_uncompressed > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
                raise CompanyIntakeValidationError(
                    "Document archive expands beyond the size limit."
                )
            if entry.file_size and entry.compress_size == 0:
                raise CompanyIntakeValidationError("Document archive compression ratio is unsafe.")
            if entry.compress_size and entry.file_size / entry.compress_size > (
                MAX_ARCHIVE_COMPRESSION_RATIO
            ):
                raise CompanyIntakeValidationError("Document archive compression ratio is unsafe.")
            names.add(normalized)
        if "[Content_Types].xml" not in names:
            raise CompanyIntakeValidationError("Office document has no content-type manifest.")
        required_prefix = "word/" if suffix == ".docx" else "xl/"
        if not any(name.startswith(required_prefix) for name in names):
            raise CompanyIntakeValidationError(
                f"Office document content does not match the {suffix} extension."
            )


def _validate_document(
    payload: bytes,
    filename: str,
    declared_mime: str | None,
) -> tuple[str, str | None]:
    if not payload:
        raise CompanyIntakeValidationError("Uploaded document is empty.")
    if len(payload) > MAX_INTAKE_BYTES:
        raise CompanyIntakeValidationError("Upload exceeds the 20 MiB local prototype limit.")
    safe_filename = _safe_filename(filename)
    suffix = Path(safe_filename).suffix.lower()
    expected_mimes = _DOCUMENT_TYPES.get(suffix)
    if expected_mimes is None:
        raise CompanyIntakeValidationError("Unsupported company document extension.")
    normalized_mime = None
    if declared_mime:
        normalized_mime = declared_mime.split(";", maxsplit=1)[0].strip().lower()
        if normalized_mime not in {*expected_mimes, "application/octet-stream"}:
            raise CompanyIntakeValidationError(
                "Declared document MIME type does not match the filename extension."
            )
    if suffix == ".pdf" and not payload.startswith(b"%PDF-"):
        raise CompanyIntakeValidationError("Document content does not match the .pdf extension.")
    if suffix == ".png" and not payload.startswith(b"\x89PNG\r\n\x1a\n"):
        raise CompanyIntakeValidationError("Document content does not match the .png extension.")
    if suffix in {".jpg", ".jpeg"} and not payload.startswith(b"\xff\xd8\xff"):
        raise CompanyIntakeValidationError("Document content does not match the image extension.")
    if suffix in {".docx", ".xlsx"}:
        _validate_zip_document(payload, suffix)
    if suffix in {".csv", ".json", ".html", ".htm", ".xml", ".txt"}:
        if b"\x00" in payload:
            raise CompanyIntakeValidationError("Text document contains binary NUL bytes.")
        try:
            payload.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise CompanyIntakeValidationError("Text document must use UTF-8 encoding.") from exc
    return safe_filename, normalized_mime


class CompanyIntakeService:
    def __init__(
        self,
        session_factory: sessionmaker[Session],
        raw_data_dir: Path,
    ) -> None:
        self._session_factory = session_factory
        self._raw_data_dir = raw_data_dir.resolve()

    def create(self, request: CompanyIntakeRequest) -> CompanyIntakeResult:
        prepared = self._prepare(request)
        created_snapshots: list[tuple[Path, str]] = []
        try:
            with self._session_factory.begin() as session:
                return self._create_prepared(session, prepared, created_snapshots)
        except Exception:
            for snapshot_path, content_sha256 in created_snapshots:
                self._remove_failed_snapshot(snapshot_path, content_sha256)
            raise

    def attach_documents(
        self,
        research_case_id: str,
        documents: tuple[CompanyDocumentUpload, ...],
        *,
        actor: str,
        classification: DataClassification,
        evidence_scope: EvidenceScope,
    ) -> tuple[CompanyIntakeResult, ...]:
        """Atomically attach one to twelve immutable documents to an existing case."""

        clean_actor, _ = _clean_actor_and_purpose(
            actor, "Attach internally sourced company evidence to the reviewed research case."
        )
        if not 1 <= len(documents) <= MAX_DOCUMENTS_PER_BATCH:
            raise CompanyIntakeValidationError("Upload between 1 and 12 documents per batch.")
        prepared: list[tuple[CompanyDocumentUpload, str, str | None, str, str]] = []
        seen_hashes: set[str] = set()
        for document in documents:
            filename, mime = _validate_document(
                document.content, document.filename, document.declared_mime
            )
            digest = sha256_bytes(document.content)
            if digest in seen_hashes:
                raise CompanyIntakeValidationError(
                    "The same document content appears more than once in this batch."
                )
            seen_hashes.add(digest)
            fingerprint = stable_hash(
                {
                    "contract": HYBRID_DOCUMENT_CONTRACT_VERSION,
                    "research_case_id": research_case_id,
                    "classification": classification.value,
                    "evidence_scope": evidence_scope.value,
                    "sha256": digest,
                }
            )
            prepared.append((document, filename, mime, digest, fingerprint))

        created_snapshots: list[tuple[Path, str]] = []
        try:
            with self._session_factory.begin() as session:
                case = session.get(ResearchCaseModel, research_case_id)
                if case is None:
                    raise CompanyIntakeValidationError("Unknown research case.")
                results: list[CompanyIntakeResult] = []
                for document, filename, mime, digest, fingerprint in prepared:
                    existing = session.scalar(
                        select(IntakeArtifactModel).where(
                            IntakeArtifactModel.fingerprint == fingerprint
                        )
                    )
                    if existing is not None:
                        results.append(
                            CompanyIntakeResult(
                                company_id=existing.company_id,
                                research_case_id=existing.research_case_id,
                                artifact_id=existing.id,
                                reused_existing=True,
                            )
                        )
                        continue
                    snapshot_path, snapshot_created = self._write_snapshot(
                        document.content, content_sha256=digest, filename=filename
                    )
                    if snapshot_created:
                        created_snapshots.append((snapshot_path, digest))
                    artifact = IntakeArtifactModel(
                        research_case_id=case.id,
                        company_id=case.company_id,
                        kind=IntakeArtifactKind.DOCUMENT.value,
                        fingerprint=fingerprint,
                        normalized_value=filename,
                        submitted_value_json={
                            "document": {
                                "filename": filename,
                                "declared_mime": mime,
                                "sha256": digest,
                            },
                            "trust_state": "untrusted",
                            "processing_boundary": "local_only",
                            "evidence_scope": evidence_scope.value,
                            "contract": HYBRID_DOCUMENT_CONTRACT_VERSION,
                        },
                        content_sha256=digest,
                        snapshot_path=str(snapshot_path),
                        original_filename=filename,
                        classification=classification.value,
                        actor=clean_actor,
                        purpose=case.purpose,
                    )
                    session.add(artifact)
                    session.flush()
                    results.append(
                        CompanyIntakeResult(
                            company_id=case.company_id,
                            research_case_id=case.id,
                            artifact_id=artifact.id,
                            reused_existing=False,
                        )
                    )
                return tuple(results)
        except Exception:
            for snapshot_path, digest in created_snapshots:
                self._remove_failed_snapshot(snapshot_path, digest)
            raise

    def propose_group_scope(
        self,
        *,
        company_id: str,
        companies_house_number: str,
        company_name: str | None,
        actor: str,
    ) -> CompanyRelationshipModel:
        """Record a held consolidated-group relationship without merging identities."""

        clean_actor, _ = _clean_actor_and_purpose(
            actor, "Propose a separately reviewed consolidated corporate group scope."
        )
        number = normalize_identifier(
            IdentifierScheme.COMPANIES_HOUSE_NUMBER, companies_house_number
        )
        if not is_valid_companies_house_number(number):
            raise CompanyIntakeValidationError(
                "Group Companies House number is not structurally valid."
            )
        clean_name = " ".join((company_name or "").split()) or f"Company {number}"
        with self._session_factory.begin() as session:
            subject = session.get(CompanyModel, company_id)
            if subject is None:
                raise CompanyIntakeValidationError("Unknown subject company.")
            subject_number = session.scalar(
                select(CompanyIdentifierModel).where(
                    CompanyIdentifierModel.company_id == subject.id,
                    CompanyIdentifierModel.scheme
                    == IdentifierScheme.COMPANIES_HOUSE_NUMBER.value,
                    CompanyIdentifierModel.normalized_value == number,
                )
            )
            if subject_number is not None:
                raise CompanyIntakeValidationError(
                    "The group scope must use a different legal entity number."
                )
            identifier = session.scalar(
                select(CompanyIdentifierModel).where(
                    CompanyIdentifierModel.scheme
                    == IdentifierScheme.COMPANIES_HOUSE_NUMBER.value,
                    CompanyIdentifierModel.normalized_value == number,
                )
            )
            if identifier is None:
                related = CompanyModel(
                    canonical_name=clean_name,
                    normalized_name=normalize_company_name(clean_name),
                    resolution_status=ResolutionStatus.UNRESOLVED.value,
                    classification=subject.classification,
                    entity_type=CompanyEntityType.REGISTERED.value,
                    jurisdiction="GB",
                    lifecycle_status=CompanyLifecycleStatus.CANDIDATE.value,
                )
                session.add(related)
                session.flush()
                identifier = CompanyIdentifierModel(
                    company_id=related.id,
                    scheme=IdentifierScheme.COMPANIES_HOUSE_NUMBER.value,
                    value=number,
                    normalized_value=number,
                    source_key=identifier_source_key(IdentifierScheme.COMPANIES_HOUSE_NUMBER),
                    reviewed=False,
                )
                session.add(identifier)
                session.flush()
            relationship = session.scalar(
                select(CompanyRelationshipModel).where(
                    CompanyRelationshipModel.subject_company_id == subject.id,
                    CompanyRelationshipModel.related_company_id == identifier.company_id,
                    CompanyRelationshipModel.relationship_type
                    == CompanyRelationshipType.CONSOLIDATED_GROUP.value,
                )
            )
            if relationship is None:
                relationship = CompanyRelationshipModel(
                    subject_company_id=subject.id,
                    related_company_id=identifier.company_id,
                    relationship_type=CompanyRelationshipType.CONSOLIDATED_GROUP.value,
                    status=LinkReviewStatus.PENDING.value,
                    proposed_by=clean_actor,
                )
                session.add(relationship)
                session.flush()
            return relationship

    def decide_group_scope(
        self,
        *,
        relationship_id: str,
        decision: IdentityDecisionType,
        actor: str,
        reason: str,
    ) -> CompanyRelationshipDecisionModel:
        clean_actor, clean_reason = _clean_actor_and_purpose(actor, reason)
        with self._session_factory.begin() as session:
            relationship = session.get(CompanyRelationshipModel, relationship_id)
            if relationship is None:
                raise CompanyIntakeValidationError("Unknown company relationship.")
            if relationship.status != LinkReviewStatus.PENDING.value:
                raise CompanyIntakeValidationError(
                    "Company relationship already has a final decision."
                )
            relationship.status = (
                LinkReviewStatus.VERIFIED.value
                if decision is IdentityDecisionType.ACCEPT
                else LinkReviewStatus.REJECTED.value
            )
            if decision is IdentityDecisionType.ACCEPT:
                related = session.get(CompanyModel, relationship.related_company_id)
                identifier = session.scalar(
                    select(CompanyIdentifierModel).where(
                        CompanyIdentifierModel.company_id == relationship.related_company_id,
                        CompanyIdentifierModel.scheme
                        == IdentifierScheme.COMPANIES_HOUSE_NUMBER.value,
                    )
                )
                if related is None or identifier is None:
                    raise CompanyIntakeValidationError(
                        "Related company identity is incomplete."
                    )
                identifier.reviewed = True
                synchronize_company_identity_review_state(session, related, accepted=True)
            record = CompanyRelationshipDecisionModel(
                company_relationship_id=relationship.id,
                decision=decision.value,
                actor=clean_actor,
                reason=clean_reason,
            )
            session.add(record)
            session.flush()
            return record

    def create_bulk(
        self,
        payload: bytes,
        *,
        filename: str,
        actor: str,
        purpose: str,
        classification: DataClassification,
    ) -> tuple[CompanyIntakeResult, ...]:
        if not payload:
            raise CompanyIntakeValidationError("Bulk intake file is empty.")
        if len(payload) > MAX_INTAKE_BYTES:
            raise CompanyIntakeValidationError("Bulk intake exceeds the 20 MiB limit.")
        rows = self._parse_bulk_rows(payload, filename)
        prepared: list[_PreparedIntake] = []
        for row_number, row in rows:
            try:
                prepared.append(
                    self._prepare(
                        CompanyIntakeRequest(
                            actor=actor,
                            purpose=purpose,
                            classification=classification,
                            companies_house_number=row.get("companies_house_number"),
                            website=row.get("website"),
                            company_name=row.get("company_name"),
                            jurisdiction=row.get("jurisdiction"),
                        ),
                        kind_override=IntakeArtifactKind.BULK_ROW,
                    )
                )
            except CompanyIntakeValidationError as exc:
                raise CompanyIntakeValidationError(f"Bulk intake row {row_number}: {exc}") from exc
        if not prepared:
            raise CompanyIntakeValidationError("Bulk intake contains no company rows.")
        created_snapshots: list[tuple[Path, str]] = []
        try:
            with self._session_factory.begin() as session:
                return tuple(
                    self._create_prepared(session, item, created_snapshots) for item in prepared
                )
        except Exception:
            for snapshot_path, content_sha256 in created_snapshots:
                self._remove_failed_snapshot(snapshot_path, content_sha256)
            raise

    def decide_identifier(
        self,
        *,
        identifier_id: str,
        decision: IdentityDecisionType,
        actor: str,
        reason: str,
    ) -> CompanyIdentifierDecisionModel:
        clean_actor, clean_reason = _clean_actor_and_purpose(actor, reason)
        with self._session_factory.begin() as session:
            identifier = session.get(CompanyIdentifierModel, identifier_id)
            if identifier is None:
                raise CompanyIntakeValidationError("Unknown company identifier.")
            company = session.get(CompanyModel, identifier.company_id)
            if company is None:
                raise CompanyIntakeValidationError("Identifier company is unavailable.")
            projection = identifier_review_projection(session, identifier)
            if projection.candidates:
                raise CompanyIntakeValidationError(
                    "Company identifier is governed by an existing identity candidate."
                )
            if projection.status != IdentityCandidateStatus.PENDING:
                raise CompanyIntakeValidationError(
                    "Company identifier already has a final decision."
                )
            identifier.reviewed = decision is IdentityDecisionType.ACCEPT
            synchronize_company_identity_review_state(
                session,
                company,
                accepted=identifier.reviewed,
            )
            record = CompanyIdentifierDecisionModel(
                company_identifier_id=identifier.id,
                decision=decision.value,
                actor=clean_actor,
                reason=clean_reason,
            )
            session.add(record)
            session.flush()
            return record

    def decide_domain(
        self,
        *,
        domain_id: str,
        decision: IdentityDecisionType,
        actor: str,
        reason: str,
    ) -> CompanyDomainDecisionModel:
        clean_actor, clean_reason = _clean_actor_and_purpose(actor, reason)
        with self._session_factory.begin() as session:
            domain = session.get(CompanyDomainModel, domain_id)
            if domain is None:
                raise CompanyIntakeValidationError("Unknown company domain.")
            if domain.status != LinkReviewStatus.PENDING.value:
                raise CompanyIntakeValidationError("Company domain already has a final decision.")
            if decision is IdentityDecisionType.ACCEPT:
                conflicting = session.scalar(
                    select(CompanyDomainModel).where(
                        CompanyDomainModel.normalized_domain == domain.normalized_domain,
                        CompanyDomainModel.status == LinkReviewStatus.VERIFIED.value,
                        CompanyDomainModel.company_id != domain.company_id,
                    )
                )
                if conflicting is not None:
                    raise CompanyIntakeValidationError(
                        "Domain is already verified for a different company."
                    )
                domain.status = LinkReviewStatus.VERIFIED.value
            else:
                domain.status = LinkReviewStatus.REJECTED.value
            record = CompanyDomainDecisionModel(
                company_domain_id=domain.id,
                decision=decision.value,
                actor=clean_actor,
                reason=clean_reason,
            )
            session.add(record)
            session.flush()
            return record

    def _prepare(
        self,
        request: CompanyIntakeRequest,
        *,
        kind_override: IntakeArtifactKind | None = None,
    ) -> _PreparedIntake:
        actor, purpose = _clean_actor_and_purpose(request.actor, request.purpose)
        number = None
        if request.companies_house_number and request.companies_house_number.strip():
            number = normalize_identifier(
                IdentifierScheme.COMPANIES_HOUSE_NUMBER,
                request.companies_house_number,
            )
            if not is_valid_companies_house_number(number):
                raise CompanyIntakeValidationError(
                    "Companies House number is not structurally valid."
                )
        website, domain = _normalise_website(request.website)
        name = " ".join(request.company_name.split()) if request.company_name else None
        if name == "":
            name = None
        if name is not None and len(name) > 255:
            raise CompanyIntakeValidationError("Company name exceeds 255 characters.")
        jurisdiction = _normalise_jurisdiction(request.jurisdiction)

        document_bytes = request.document_bytes
        safe_filename = None
        content_sha256 = None
        if document_bytes is not None:
            if name is None:
                raise CompanyIntakeValidationError(
                    "Document intake requires a declared company name."
                )
            if request.document_filename is None:
                raise CompanyIntakeValidationError("Document intake requires a filename.")
            safe_filename, normalized_mime = _validate_document(
                document_bytes,
                request.document_filename,
                request.declared_mime,
            )
            content_sha256 = sha256_bytes(document_bytes)
        else:
            normalized_mime = None
        if number is None and website is None and name is None and document_bytes is None:
            raise CompanyIntakeValidationError(
                "Company intake requires at least one identity input."
            )
        if jurisdiction is not None and name is None and number is None:
            raise CompanyIntakeValidationError(
                "Jurisdiction requires a company name or Companies House number."
            )

        if kind_override is not None:
            kind = kind_override
        elif number is not None:
            kind = IntakeArtifactKind.COMPANIES_HOUSE_NUMBER
        elif website is not None:
            kind = IntakeArtifactKind.WEBSITE
        elif document_bytes is not None:
            kind = IntakeArtifactKind.DOCUMENT
        else:
            kind = IntakeArtifactKind.NAME_JURISDICTION
        submitted_value: dict[str, Any] = {
            "companies_house_number": number,
            "website": website,
            "company_name": name,
            "jurisdiction": jurisdiction,
            "document": (
                {
                    "filename": safe_filename,
                    "declared_mime": normalized_mime,
                    "sha256": content_sha256,
                }
                if document_bytes is not None
                else None
            ),
            "trust_state": "untrusted" if document_bytes is not None else "submitted_claim",
        }
        fingerprint = stable_hash(
            {
                "contract": INTAKE_CONTRACT_VERSION,
                "template_version": CORE_TEMPLATE_VERSION,
                "kind": kind.value,
                "purpose": purpose,
                "classification": request.classification.value,
                "submitted_value": submitted_value,
            }
        )
        return _PreparedIntake(
            kind=kind,
            actor=actor,
            purpose=purpose,
            classification=request.classification,
            companies_house_number=number,
            website=website,
            domain=domain,
            company_name=name,
            jurisdiction=jurisdiction,
            document_bytes=document_bytes,
            document_filename=safe_filename,
            declared_mime=normalized_mime,
            content_sha256=content_sha256,
            submitted_value=submitted_value,
            fingerprint=fingerprint,
        )

    def _create_prepared(
        self,
        session: Session,
        prepared: _PreparedIntake,
        created_snapshots: list[tuple[Path, str]],
    ) -> CompanyIntakeResult:
        existing = session.scalar(
            select(IntakeArtifactModel).where(
                IntakeArtifactModel.fingerprint == prepared.fingerprint
            )
        )
        if existing is not None:
            return CompanyIntakeResult(
                company_id=existing.company_id,
                research_case_id=existing.research_case_id,
                artifact_id=existing.id,
                reused_existing=True,
            )

        template_version = self._ensure_core_template(session)
        company = self._select_or_create_company(session, prepared)
        case = session.scalar(
            select(ResearchCaseModel).where(
                ResearchCaseModel.company_id == company.id,
                ResearchCaseModel.purpose == prepared.purpose,
                ResearchCaseModel.template_version_id == template_version.id,
                ResearchCaseModel.classification == prepared.classification.value,
            )
        )
        if case is None:
            case = ResearchCaseModel(
                company_id=company.id,
                template_version_id=template_version.id,
                purpose=prepared.purpose,
                classification=prepared.classification.value,
                status=(
                    ResearchCaseStatus.READY.value
                    if company.resolution_status == ResolutionStatus.RESOLVED.value
                    else ResearchCaseStatus.IDENTITY_HOLD.value
                ),
                created_by=prepared.actor,
            )
            session.add(case)
            session.flush()

        snapshot_path = None
        if prepared.document_bytes is not None:
            assert prepared.content_sha256 is not None
            assert prepared.document_filename is not None
            snapshot_path, snapshot_created = self._write_snapshot(
                prepared.document_bytes,
                content_sha256=prepared.content_sha256,
                filename=prepared.document_filename,
            )
            if snapshot_created:
                created_snapshots.append((snapshot_path, prepared.content_sha256))
        artifact = IntakeArtifactModel(
            research_case_id=case.id,
            company_id=company.id,
            kind=prepared.kind.value,
            fingerprint=prepared.fingerprint,
            normalized_value=(
                prepared.companies_house_number or prepared.website or prepared.company_name
            ),
            submitted_value_json=prepared.submitted_value,
            content_sha256=prepared.content_sha256,
            snapshot_path=str(snapshot_path) if snapshot_path else None,
            original_filename=prepared.document_filename,
            classification=prepared.classification.value,
            actor=prepared.actor,
            purpose=prepared.purpose,
        )
        session.add(artifact)
        session.flush()
        return CompanyIntakeResult(
            company_id=company.id,
            research_case_id=case.id,
            artifact_id=artifact.id,
            reused_existing=False,
        )

    @staticmethod
    def _ensure_core_template(session: Session) -> ResearchTemplateVersionModel:
        template = session.scalar(
            select(ResearchTemplateModel).where(ResearchTemplateModel.key == CORE_TEMPLATE_KEY)
        )
        if template is None:
            template = ResearchTemplateModel(
                key=CORE_TEMPLATE_KEY,
                name="Core company profile",
            )
            session.add(template)
            session.flush()
        version = session.scalar(
            select(ResearchTemplateVersionModel).where(
                ResearchTemplateVersionModel.template_id == template.id,
                ResearchTemplateVersionModel.version == CORE_TEMPLATE_VERSION,
            )
        )
        if version is not None:
            persisted_contract = {
                "key": template.key,
                "version": version.version,
                "objective": version.objective,
                "required_capabilities": version.required_capabilities_json,
                "optional_capabilities": version.optional_capabilities_json,
                "claim_keys": version.claim_keys_json,
                "budgets": version.budgets_json,
            }
            if (
                version.sha256 != CORE_TEMPLATE_SHA256
                or stable_hash(persisted_contract) != CORE_TEMPLATE_SHA256
            ):
                raise RuntimeError("Core company-profile template version has drifted.")
            return version
        version = ResearchTemplateVersionModel(
            template_id=template.id,
            version=CORE_TEMPLATE_VERSION,
            objective=_CORE_TEMPLATE_CONTRACT["objective"],
            required_capabilities_json=_CORE_TEMPLATE_CONTRACT["required_capabilities"],
            optional_capabilities_json=_CORE_TEMPLATE_CONTRACT["optional_capabilities"],
            claim_keys_json=_CORE_TEMPLATE_CONTRACT["claim_keys"],
            budgets_json=_CORE_TEMPLATE_CONTRACT["budgets"],
            sha256=CORE_TEMPLATE_SHA256,
        )
        session.add(version)
        session.flush()
        return version

    @staticmethod
    def _select_or_create_company(
        session: Session,
        prepared: _PreparedIntake,
    ) -> CompanyModel:
        company = None
        if prepared.companies_house_number is not None:
            identifier = session.scalar(
                select(CompanyIdentifierModel).where(
                    CompanyIdentifierModel.scheme == IdentifierScheme.COMPANIES_HOUSE_NUMBER.value,
                    CompanyIdentifierModel.normalized_value == prepared.companies_house_number,
                )
            )
            if identifier is not None:
                company = session.get(CompanyModel, identifier.company_id)
                if company is None:
                    raise RuntimeError("Company identifier references a missing company.")
                if company.classification != prepared.classification.value:
                    raise CompanyIntakeValidationError(
                        "Exact identifier already belongs to a different classification boundary."
                    )
        if company is None:
            label = prepared.company_name
            if label is None and prepared.companies_house_number is not None:
                label = f"Unresolved company (CH {prepared.companies_house_number})"
            if label is None and prepared.domain is not None:
                label = f"Unresolved company ({prepared.domain})"
            assert label is not None
            company = CompanyModel(
                canonical_name=label,
                normalized_name=normalize_company_name(label),
                external_id=None,
                resolution_status=ResolutionStatus.UNRESOLVED.value,
                classification=prepared.classification.value,
                entity_type=(
                    CompanyEntityType.REGISTERED.value
                    if prepared.companies_house_number is not None
                    else CompanyEntityType.UNKNOWN.value
                ),
                jurisdiction=prepared.jurisdiction,
                lifecycle_status=CompanyLifecycleStatus.CANDIDATE.value,
            )
            session.add(company)
            session.flush()

        if prepared.companies_house_number is not None:
            identifier = session.scalar(
                select(CompanyIdentifierModel).where(
                    CompanyIdentifierModel.scheme == IdentifierScheme.COMPANIES_HOUSE_NUMBER.value,
                    CompanyIdentifierModel.normalized_value == prepared.companies_house_number,
                )
            )
            if identifier is None:
                session.add(
                    CompanyIdentifierModel(
                        company_id=company.id,
                        scheme=IdentifierScheme.COMPANIES_HOUSE_NUMBER.value,
                        value=prepared.companies_house_number,
                        normalized_value=prepared.companies_house_number,
                        source_key=identifier_source_key(IdentifierScheme.COMPANIES_HOUSE_NUMBER),
                        reviewed=False,
                    )
                )
        if prepared.domain is not None and prepared.website is not None:
            domain = session.scalar(
                select(CompanyDomainModel).where(
                    CompanyDomainModel.company_id == company.id,
                    CompanyDomainModel.normalized_domain == prepared.domain,
                )
            )
            if domain is None:
                session.add(
                    CompanyDomainModel(
                        company_id=company.id,
                        url=prepared.website,
                        normalized_domain=prepared.domain,
                        status=LinkReviewStatus.PENDING.value,
                    )
                )
        session.flush()
        return company

    def _write_snapshot(
        self,
        payload: bytes,
        *,
        content_sha256: str,
        filename: str,
    ) -> tuple[Path, bool]:
        target_dir = self._raw_data_dir / "company-intakes" / content_sha256
        target_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
        target_dir.chmod(0o700)
        target = target_dir / filename
        if target.exists():
            if not target.is_file() or sha256_bytes(target.read_bytes()) != content_sha256:
                raise RuntimeError("Existing company-intake snapshot does not match its checksum.")
            target.chmod(0o600)
            return target, False
        try:
            handle = target.open("xb")
        except FileExistsError:
            if sha256_bytes(target.read_bytes()) != content_sha256:
                raise RuntimeError("Concurrent intake snapshot content disagrees.") from None
            target.chmod(0o600)
            return target, False
        try:
            with handle:
                handle.write(payload)
                handle.flush()
                os.fsync(handle.fileno())
            target.chmod(0o600)
        except BaseException:
            with suppress(OSError):
                target.unlink()
            with suppress(OSError):
                target.parent.rmdir()
            raise
        return target, True

    @staticmethod
    def _remove_failed_snapshot(path: Path, expected_hash: str) -> None:
        if path.is_file() and sha256_bytes(path.read_bytes()) == expected_hash:
            path.unlink()
            with suppress(OSError):
                path.parent.rmdir()

    @staticmethod
    def _parse_bulk_rows(payload: bytes, filename: str) -> tuple[tuple[int, dict[str, str]], ...]:
        suffix = Path(filename).suffix.lower()
        rows: list[tuple[int, dict[str, str]]] = []
        required = {
            "companies_house_number",
            "website",
            "company_name",
            "jurisdiction",
        }
        if suffix == ".csv":
            try:
                decoded = payload.decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise CompanyIntakeValidationError("Bulk CSV must use UTF-8 encoding.") from exc
            reader = csv.DictReader(io.StringIO(decoded))
            csv_headers = set(reader.fieldnames or ())
            if len(csv_headers) > MAX_BULK_COLUMNS:
                raise CompanyIntakeValidationError("Bulk CSV exceeds the column limit.")
            if not required <= csv_headers:
                raise CompanyIntakeValidationError(
                    "Bulk intake requires companies_house_number, website, company_name, and "
                    "jurisdiction columns."
                )
            for row_number, row in enumerate(reader, start=2):
                if row_number > MAX_BULK_ROWS + 1:
                    raise CompanyIntakeValidationError("Bulk CSV exceeds the row limit.")
                values = {key: (row.get(key) or "").strip() for key in required}
                if any(values.values()):
                    rows.append((row_number, values))
            return tuple(rows)
        if suffix != ".xlsx":
            raise CompanyIntakeValidationError("Bulk intake supports only CSV or XLSX files.")
        _validate_zip_document(payload, suffix)
        try:
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
        except Exception as exc:
            raise CompanyIntakeValidationError("Bulk XLSX is malformed or unreadable.") from exc
        try:
            sheet = workbook.active
            if sheet.max_row > MAX_BULK_ROWS + 1 or sheet.max_column > MAX_BULK_COLUMNS:
                raise CompanyIntakeValidationError("Bulk XLSX exceeds the row or column limit.")
            row_iterator = sheet.iter_rows()
            header_row = next(row_iterator, None)
            if header_row is None:
                return ()
            xlsx_headers = [str(cell.value or "").strip() for cell in header_row]
            if not required <= set(xlsx_headers):
                raise CompanyIntakeValidationError(
                    "Bulk intake requires companies_house_number, website, company_name, and "
                    "jurisdiction columns."
                )
            indices = {key: xlsx_headers.index(key) for key in required}
            for row_number, row in enumerate(row_iterator, start=2):
                if row_number > MAX_BULK_ROWS + 1:
                    raise CompanyIntakeValidationError("Bulk XLSX exceeds the row limit.")
                if any(cell.data_type == "f" for cell in row):
                    raise CompanyIntakeValidationError(
                        f"Bulk intake row {row_number} contains a formula."
                    )
                values = {
                    key: (
                        str(row[index].value).strip()
                        if index < len(row) and row[index].value is not None
                        else ""
                    )
                    for key, index in indices.items()
                }
                if any(values.values()):
                    rows.append((row_number, values))
            return tuple(rows)
        finally:
            workbook.close()
