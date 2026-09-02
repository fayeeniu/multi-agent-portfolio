"""Bounded local text and CBIT extraction for authorised intake documents.

Document bytes never cross the external-model boundary. Extraction produces only
exact spans from a deterministic local text projection; unsupported or ambiguous
content remains inventory rather than evidence.
"""

from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader

from .cbit_contract import CBIT_ROWS, CbitRowRole, CbitValueShape
from .enums import ResearchClaimCategory

MAX_DOCUMENT_PAGES = 300
MAX_DOCUMENT_TEXT_CHARS = 500_000


@dataclass(frozen=True, slots=True)
class LocalMetricSpan:
    category: ResearchClaimCategory
    subject_key: str
    evidence_span: str
    value: str


_CATEGORY_MAP = {
    "Employment and economic impact": ResearchClaimCategory.PERFORMANCE,
    "Research and development": ResearchClaimCategory.TECHNOLOGY,
    "Technology readiness": ResearchClaimCategory.TECHNOLOGY,
    "Products and processes": ResearchClaimCategory.PRODUCTS_MARKET,
    "Funding and investments": ResearchClaimCategory.FUNDING,
    "Market and partnerships": ResearchClaimCategory.PRODUCTS_MARKET,
    "Financial impact": ResearchClaimCategory.PERFORMANCE,
    "Diversity and sustainability": ResearchClaimCategory.OTHER,
    "Policy and influence": ResearchClaimCategory.REGULATION,
    "AI operational efficiency": ResearchClaimCategory.PERFORMANCE,
    "AI adoption and readiness": ResearchClaimCategory.TECHNOLOGY,
}


def _bounded(lines: list[str]) -> str:
    clean = [re.sub(r"\s+", " ", line).strip() for line in lines]
    return "\n".join(line for line in clean if line)[:MAX_DOCUMENT_TEXT_CHARS]


def _docx_text(payload: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        root = ElementTree.fromstring(archive.read("word/document.xml"))
    lines: list[str] = []
    for row in root.iter():
        if row.tag.rsplit("}", 1)[-1] != "tr":
            continue
        cells = []
        for cell in row:
            if cell.tag.rsplit("}", 1)[-1] != "tc":
                continue
            cells.append(" ".join(text.text or "" for text in cell.iter() if text.tag.rsplit("}", 1)[-1] == "t"))
        if cells:
            lines.append(" | ".join(cells))
    if not lines:
        lines = [text.text or "" for text in root.iter() if text.tag.rsplit("}", 1)[-1] == "t"]
    return _bounded(lines)


def _xlsx_text(payload: bytes) -> str:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    try:
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"Sheet | {sheet.title}")
            for row in sheet.iter_rows():
                if any(cell.data_type == "f" for cell in row):
                    continue
                values = [str(cell.value).strip() for cell in row if cell.value is not None]
                if values:
                    lines.append(" | ".join(values))
                if sum(len(line) for line in lines) >= MAX_DOCUMENT_TEXT_CHARS:
                    break
        return _bounded(lines)
    finally:
        workbook.close()


def document_text(payload: bytes, filename: str) -> str:
    """Return a bounded deterministic local text projection for a supported document."""

    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        reader = PdfReader(io.BytesIO(payload), strict=True)
        if reader.is_encrypted or len(reader.pages) > MAX_DOCUMENT_PAGES:
            return ""
        return _bounded([(page.extract_text() or "") for page in reader.pages])
    if suffix == ".docx":
        return _docx_text(payload)
    if suffix == ".xlsx":
        return _xlsx_text(payload)
    if suffix == ".csv":
        rows = csv.reader(io.StringIO(payload.decode("utf-8-sig")))
        return _bounded([" | ".join(cell.strip() for cell in row) for row in rows])
    if suffix == ".json":
        parsed = json.loads(payload.decode("utf-8-sig"))
        return _bounded(json.dumps(parsed, ensure_ascii=False, indent=2).splitlines())
    if suffix in {".html", ".htm", ".xml", ".txt"}:
        return _bounded(payload.decode("utf-8-sig").splitlines())
    return ""


def extract_cbit_spans(text: str) -> tuple[LocalMetricSpan, ...]:
    """Extract exact label/value rows; never infer counts, periods, or semantics."""

    spans: list[LocalMetricSpan] = []
    seen: set[str] = set()
    for line in text.splitlines():
        clean = re.sub(r"\s+", " ", line).strip()
        if not clean:
            continue
        for row in CBIT_ROWS:
            if row.role is not CbitRowRole.INPUT or row.metric_key is None:
                continue
            match = re.match(rf"^{re.escape(row.label)}\s*(?:\||:|\t|–|—|-)+\s*(.+)$", clean, re.I)
            if match is None:
                continue
            value = match.group(1).strip()
            if not value or value.casefold() in {"n/a", "na", "not available", "unknown"}:
                continue
            if row.value_shape in {
                CbitValueShape.INTEGER,
                CbitValueShape.CURRENCY,
                CbitValueShape.PERCENTAGE,
                CbitValueShape.REPORTED_DURATION,
            } and re.search(r"\d", value) is None:
                continue
            if row.metric_key in seen:
                continue
            seen.add(row.metric_key)
            spans.append(
                LocalMetricSpan(
                    category=_CATEGORY_MAP.get(row.category, ResearchClaimCategory.OTHER),
                    subject_key=row.metric_key,
                    evidence_span=clean,
                    value=value[:100],
                )
            )
    return tuple(spans)
